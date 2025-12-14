import io
import re
import math
from typing import Dict, List, Tuple, Any

import pandas as pd
from openpyxl import load_workbook


# ---------------------------
# Helpers
# ---------------------------

def _norm(v) -> str:
    """Normalize values for reliable matching"""
    if v is None:
        return ""

    try:
        if isinstance(v, float) and math.isnan(v):
            return ""
    except Exception:
        pass

    if isinstance(v, (int, float)):
        if isinstance(v, float) and v.is_integer():
            v = int(v)
        return str(v).strip().lower()

    s = str(v).strip().lower()
    s = re.sub(r"\s+", " ", s)

    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    return s


def _find_jot_col(df, patterns):
    """Finds a column in the Jotform DataFrame matching a set of keywords."""
    for col in df.columns:
        col_u = str(col).strip().upper().replace("'", "")
        for p in patterns:
            if all(word in col_u for word in p):
                return col
    return None


def _find_sheet_name(wb, keywords, exact_match=False):
    """Finds a sheet name."""
    for name in wb.sheetnames:
        name_u = name.upper()
        if exact_match and keywords:
            if name_u == keywords[0].upper():
                return name
        if not exact_match:
            if any(k in name_u for k in keywords):
                return name
    return None


def _find_header_map(ws) -> Dict[str, int]:
    """Finds headers mapping (Upper -> Index)."""
    MAX_HEADER_SEARCH_ROWS = 10
    for r in range(1, min(ws.max_row + 1, MAX_HEADER_SEARCH_ROWS + 1)):
        headers = {}
        found_header_count = 0
        for c in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=r, column=c).value
            if cell_value is not None:
                headers[str(cell_value).strip().upper()] = c
                found_header_count += 1
        
        if found_header_count > 0:
            return headers
    return {}


def _parse_percentage(val):
    """Parses a percentage value (e.g. 50, '50%', 0.5) into a float 0-100."""
    if val is None:
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace("%", "").strip()
        v = float(val)
        # If value is small (e.g. 0.5), assume it means 50%
        if v <= 1.0 and v > 0:
            return v * 100
        return v
    except:
        return 0.0

def _extract_number(val):
    """
    Extracts the first numeric value from a string.
    Example: 'EEP-50' -> 50.0, '50%' -> 50.0
    """
    if val is None:
        return None
    
    s = str(val).strip()
    if not s:
        return None
        
    match = re.search(r"(\d+(\.\d+)?)", s)
    if match:
        try:
            return float(match.group(1))
        except:
            return None
    return None


def _map_capacity(cap_val, is_publisher=False):
    """Maps capacity codes to full text."""
    if not cap_val:
        return ""
    
    s = str(cap_val).strip().upper()
    
    # Publisher Mapping
    if is_publisher:
        if s == "OP":
            return "Original Publisher"
        return s
        
    # Composer Mapping
    if s == "C":
        return "Composer"
    if s == "A":
        return "Lyrics"
    if s in ["AC", "CA"]:
        return "Lyrics and Music"
    return s


# ---------------------------
# Main processor
# ---------------------------

def process_alternate_titles(curve_reexport_buffer, jotform_file_buffer) -> io.BytesIO | str:
    try:
        # ---- Load Jotform ----
        jot = pd.read_excel(jotform_file_buffer, header=0)

        # Basic Jotform Columns for Alt Titles
        jot_title_col = _find_jot_col(jot, [["TITLE"]])
        jot_eep_col = _find_jot_col(jot, [["EEP", "MASTER", "CATALOG", "NUMBER"], ["EEP", "CATALOG", "NUMBER"]])
        jot_alt_col = _find_jot_col(jot, [["ALTERNATE", "TITLE"], ["ALT", "TITLE"], ["ALTERNATE"]])
        
        # Additional Jotform Columns for IP Chain
        jot_writer_total_col = _find_jot_col(jot, [["WRITER", "TOTAL"], ["TOTAL", "WRITERS"]])
        
        # Build Lookups
        jot_alt_lookup: Dict[Tuple[str, str], List[str]] = {}
        jot_row_lookup: Dict[str, Any] = {} # Lookup by EEP Master Number -> Row Data

        # Normalize Jotform Data for Lookup
        for idx, row in jot.iterrows():
            # 1. Alt Title Lookup Key (Title + EEP)
            title_val = row.get(jot_title_col)
            eep_val = row.get(jot_eep_col)
            
            if eep_val is not None:
                eep_norm = _norm(eep_val)
                jot_row_lookup[eep_norm] = row # Store full row for IP Chain lookup
                
                if title_val is not None:
                    title_norm = _norm(title_val)
                    if title_norm:
                        # Process Alt Titles
                        if jot_alt_col:
                            alt_raw = row.get(jot_alt_col)
                            if isinstance(alt_raw, str) and alt_raw.strip():
                                alt_list = [a.strip() for a in alt_raw.replace("\r", "\n").split("\n") if a.strip()]
                                if alt_list:
                                    jot_alt_lookup[(title_norm, eep_norm)] = alt_list


        # ---- Load Curve re-export ----
        curve_reexport_buffer.seek(0)
        wb = load_workbook(curve_reexport_buffer)

        works_sheet_name = _find_sheet_name(wb, ["WORK"])
        alt_sheet_name = _find_sheet_name(wb, ["ALTERNATE"])
        ip_sheet_name = _find_sheet_name(wb, ["IP Chain"], exact_match=True)

        if not works_sheet_name: raise ValueError("Curve file: 'Works' sheet not found")
        if not alt_sheet_name: raise ValueError("Curve file: 'Alternate Titles' sheet not found")
        if not ip_sheet_name: raise ValueError("Curve file: 'IP Chain' sheet not found")

        ws_works = wb[works_sheet_name]
        ws_alt = wb[alt_sheet_name]
        ws_ip = wb[ip_sheet_name]

        # ---- Map Headers ----
        works_headers = _find_header_map(ws_works)
        alt_headers = _find_header_map(ws_alt)
        ip_headers = _find_header_map(ws_ip)

        # Helpers for source cols
        def wh(keywords):
            for k, c in works_headers.items():
                if all(w in k for w in keywords): return c
            return None
            
        def iph(keywords):
            for k, c in ip_headers.items():
                if all(w in k for w in keywords): return c
            return None

        # --- Works Sheet Columns ---
        col_work_id = wh(["ID"])
        col_title = wh(["TITLE"])
        col_main_id = wh(["MAIN", "IDENTIFIER"])
        col_tunecode = wh(["TUNECODE"])
        col_foreign = wh(["FOREIGN", "ID"])
        col_lang = wh(["LANGUAGE"])
        col_territory = wh(["TERRITORIES"])

        if not col_work_id: raise ValueError("Works sheet missing 'ID' column")
        if not col_foreign: raise ValueError("Works sheet missing 'Foreign ID' column (needed for lookup)")

        # --- Dynamic IP Chain Column Finder (Participant 1..10) ---
        participant_cols = {}
        for i in range(1, 11): # Max 10
            p_map = {}
            prefix = f"PARTICIPANT {i}"
            
            p_map['TYPE'] = iph([prefix, "TYPE"])
            p_map['NAME'] = iph([prefix, "NAME"])
            p_map['FIRST'] = iph([prefix, "FIRST"])
            p_map['MIDDLE'] = iph([prefix, "MIDDLE"])
            p_map['SURNAME'] = iph([prefix, "SURNAME"])
            p_map['CAE'] = iph([prefix, "CAE"])
            p_map['CONTROLLED'] = iph([prefix, "CONTROLLED"])
            p_map['MECH_OWN'] = iph([prefix, "MECHANICAL", "OWNED"])
            p_map['MECH_COLL'] = iph([prefix, "MECHANICAL", "COLLECTED"])
            p_map['PERF_OWN'] = iph([prefix, "PERFORMANCE", "OWNED"])
            p_map['PERF_COLL'] = iph([prefix, "PERFORMANCE", "COLLECTED"])
            p_map['CAPACITY'] = iph([prefix, "CAPACITY"])
            
            participant_cols[i] = p_map


        # ---- Clean Target Sheets ----
        # Clear Alternate Titles data from row 3 onwards (keep header + instruction row)
        if ws_alt.max_row > 2:
            ws_alt.delete_rows(3, ws_alt.max_row - 2)
            
        # Clear IP Chain data from row 3 onwards (keep rows 1 and 2 only)
        if ws_ip.max_row > 2:
            ws_ip.delete_rows(3, ws_ip.max_row - 2)

        write_row_alt = 3
        write_row_ip = 3 

        # ---------------------------
        # Processing Loops
        # ---------------------------

        for r in range(2, ws_works.max_row + 1):
            # Source Data
            w_work_id_val = ws_works.cell(r, col_work_id).value
            if w_work_id_val is None: continue

            w_title_val = ws_works.cell(r, col_title).value
            w_foreign_val = ws_works.cell(r, col_foreign).value
            w_main_id_val = ws_works.cell(r, col_main_id).value if col_main_id else None
            w_tunecode_val = ws_works.cell(r, col_tunecode).value if col_tunecode else None
            w_lang_val = ws_works.cell(r, col_lang).value if col_lang else None
            w_terr_val = ws_works.cell(r, col_territory).value if col_territory else None

            w_title_norm = _norm(w_title_val)
            w_foreign_norm = _norm(w_foreign_val)

            # ---------------------------
            # 1. IP CHAIN LOGIC
            # ---------------------------
            
            matched_jot_row = jot_row_lookup.get(w_foreign_norm)
            
            # Helper to write common work data
            def write_ip_work_info(target_row):
                ws_ip.cell(target_row, participant_cols[1]['TYPE'] or 1).value = None 
                if iph(["WORK", "ID"]): ws_ip.cell(target_row, iph(["WORK", "ID"])).value = w_work_id_val
                if iph(["WORK", "TITLE"]): ws_ip.cell(target_row, iph(["WORK", "TITLE"])).value = w_title_val
                if iph(["WORK", "MAIN"]): ws_ip.cell(target_row, iph(["WORK", "MAIN"])).value = w_main_id_val
                if iph(["WORK", "TUNE"]): ws_ip.cell(target_row, iph(["WORK", "TUNE"])).value = w_tunecode_val
                if iph(["TERRITORY"]): ws_ip.cell(target_row, iph(["TERRITORY"])).value = w_terr_val

            if matched_jot_row is not None:
                # 1. Get Writer Total
                writer_total = 0
                if jot_writer_total_col:
                    try:
                        writer_total = int(float(matched_jot_row.get(jot_writer_total_col, 0)))
                    except:
                        writer_total = 0
                
                # 2. Extract Data for all writers
                extracted_writers = []
                for i in range(1, min(writer_total + 1, 11)):
                    # Dynamic Columns
                    c_first_col = _find_jot_col(jot, [[f"COMPOSER {i} FIRST"]])
                    c_mid_col = _find_jot_col(jot, [[f"COMPOSER {i} MIDDLE"]])
                    c_last_col = _find_jot_col(jot, [[f"COMPOSER {i} LAST"]])
                    c_share_col = _find_jot_col(jot, [[f"COMPOSER {i} SHARE"]])
                    c_cap_col = _find_jot_col(jot, [[f"COMPOSER {i} CAPACITY"]])
                    c_cae_col = _find_jot_col(jot, [[f"COMPOSER {i} CAE"]])
                    c_controlled_col = _find_jot_col(jot, [[f"COMPOSER {i} CONTROLLED"]]) # <--- NEW: Composer Controlled Column
                    
                    p_name_col = _find_jot_col(jot, [[f"PUBLISHER {i} NAME"]])
                    p_cae_col = _find_jot_col(jot, [[f"PUBLISHER {i} CAE"]])
                    p_cap_col = _find_jot_col(jot, [[f"PUBLISHER {i} CAPACITY"]])
                    
                    eep_share_col = _find_jot_col(jot, [["ELITE", "EMBASSY", "REPRESENTS", "%"]])

                    # Values
                    c_first = matched_jot_row.get(c_first_col) if c_first_col else ""
                    c_mid = matched_jot_row.get(c_mid_col) if c_mid_col else ""
                    c_last = matched_jot_row.get(c_last_col) if c_last_col else ""
                    
                    c_first = str(c_first).strip() if c_first else ""
                    c_mid = str(c_mid).strip() if c_mid else ""
                    c_last = str(c_last).strip() if c_last else ""
                    
                    full_name = f"{c_first} {c_mid} {c_last}".replace("  ", " ").strip()
                    
                    c_share = _parse_percentage(matched_jot_row.get(c_share_col))
                    c_cap = _map_capacity(matched_jot_row.get(c_cap_col), is_publisher=False)
                    c_cae = matched_jot_row.get(c_cae_col)
                    
                    # Composer Controlled Status (Y/N -> True/False)
                    c_controlled_val = str(matched_jot_row.get(c_controlled_col)).strip().upper() if c_controlled_col else ""
                    is_composer_controlled = (c_controlled_val == "Y")
                    
                    p_name = matched_jot_row.get(p_name_col)
                    p_name_str = str(p_name).strip() if p_name else ""
                    p_cae = matched_jot_row.get(p_cae_col)
                    p_cap = _map_capacity(matched_jot_row.get(p_cap_col), is_publisher=True)

                    # Mechanical % for Elite
                    raw_mech = matched_jot_row.get(eep_share_col) if eep_share_col else None
                    mech_val = _extract_number(raw_mech) # e.g., 50.0

                    # Check Publisher Controlled Status (used for grouping rows)
                    norm_p_name = _norm(p_name_str)
                    is_publisher_controlled = "elite embassy publishing" in norm_p_name or "music embassies publishing" in norm_p_name
                    
                    w_obj = {
                        "c_first": c_first,
                        "c_mid": c_mid,
                        "c_last": c_last,
                        "c_name": full_name,
                        "c_share": c_share,
                        "c_cap": c_cap,
                        "c_cae": c_cae,
                        "p_name": p_name_str,
                        "p_cae": p_cae,
                        "p_cap": p_cap,
                        "mech_val": mech_val if mech_val is not None else 0.0,
                        "is_controlled": is_composer_controlled, # <--- Composer's specific status (Y/N)
                        "is_publisher_controlled": is_publisher_controlled # <--- Publisher's status (used for grouping)
                    }
                    extracted_writers.append(w_obj)

                # 3. Sort/Group Writers
                # Grouping is still based on the publisher's controlled status
                controlled_group = [w for w in extracted_writers if w['is_publisher_controlled']]
                other_group = [w for w in extracted_writers if not w['is_publisher_controlled']]

                # Determine the Elite Mechanical Value (Used for calculating Copyright Control remainder)
                elite_mech_value = 0.0
                for w in controlled_group:
                    if w['mech_val'] > 0:
                        elite_mech_value = w['mech_val']
                        break
                if elite_mech_value == 0.0:
                    for w in extracted_writers:
                        if w['mech_val'] > 0:
                            elite_mech_value = w['mech_val']
                            break

                rows_to_create = []
                if controlled_group:
                    rows_to_create.append({"type": "controlled", "writers": controlled_group})
                if other_group:
                    rows_to_create.append({"type": "other", "writers": other_group})
                
                # 4. Generate Rows
                for row_data in rows_to_create:
                    r_type = row_data['type']
                    writers = row_data['writers']
                    
                    write_ip_work_info(write_row_ip)
                    
                    # --- Participant 1: Publisher ---
                    p1_cols = participant_cols[1]
                    
                    # Determine P1 values
                    if r_type == "controlled":
                        # Controlled: Use extracted value
                        p1_name = writers[0]['p_name']
                        p1_cae = writers[0]['p_cae']
                        p1_is_ctrl = True
                        p1_mech = elite_mech_value 
                        p1_cap = writers[0]['p_cap']
                    else:
                        # Copyright Control: Use 100 - Elite Value
                        p1_name = "Copyright Control"
                        p1_cae = "00000000000" # Fixed CAE for CC
                        p1_is_ctrl = False
                        p1_mech = 100.0 - elite_mech_value 
                        p1_cap = "Original Publisher" 

                    # Performance is always 50% of Owned Mech
                    p1_perf = p1_mech * 0.5

                    # Write P1
                    if p1_cols['TYPE']: ws_ip.cell(write_row_ip, p1_cols['TYPE']).value = "Publisher"
                    if p1_cols['NAME']: ws_ip.cell(write_row_ip, p1_cols['NAME']).value = p1_name
                    # P1 Name parts blank per instructions
                    if p1_cols['CAE']: ws_ip.cell(write_row_ip, p1_cols['CAE']).value = p1_cae
                    if p1_cols['CONTROLLED']: ws_ip.cell(write_row_ip, p1_cols['CONTROLLED']).value = p1_is_ctrl
                    if p1_cols['CAPACITY']: ws_ip.cell(write_row_ip, p1_cols['CAPACITY']).value = p1_cap
                    
                    if p1_cols['MECH_OWN']: ws_ip.cell(write_row_ip, p1_cols['MECH_OWN']).value = p1_mech
                    if p1_cols['MECH_COLL']: ws_ip.cell(write_row_ip, p1_cols['MECH_COLL']).value = p1_mech
                    if p1_cols['PERF_OWN']: ws_ip.cell(write_row_ip, p1_cols['PERF_OWN']).value = p1_perf
                    if p1_cols['PERF_COLL']: ws_ip.cell(write_row_ip, p1_cols['PERF_COLL']).value = p1_perf

                    # --- Participant 2..N+1: Composers ---
                    for i, w in enumerate(writers):
                        p_idx = i + 2 # Start at Participant 2
                        if p_idx > 10: break

                        p_cols = participant_cols[p_idx]
                        
                        # Calc values
                        comp_perf = w['c_share'] * 0.5
                        
                        if p_cols['TYPE']: ws_ip.cell(write_row_ip, p_cols['TYPE']).value = "Composer"
                        if p_cols['NAME']: ws_ip.cell(write_row_ip, p_cols['NAME']).value = w['c_name']
                        if p_cols['FIRST']: ws_ip.cell(write_row_ip, p_cols['FIRST']).value = w['c_first']
                        if p_cols['MIDDLE']: ws_ip.cell(write_row_ip, p_cols['MIDDLE']).value = w['c_mid']
                        if p_cols['SURNAME']: ws_ip.cell(write_row_ip, p_cols['SURNAME']).value = w['c_last']
                        if p_cols['CAE']: ws_ip.cell(write_row_ip, p_cols['CAE']).value = w['c_cae']
                        
                        # Controlled: Uses the Composer's specific status (Y/N -> True/False)
                        if p_cols['CONTROLLED']: ws_ip.cell(write_row_ip, p_cols['CONTROLLED']).value = w['is_controlled']
                        
                        if p_cols['CAPACITY']: ws_ip.cell(write_row_ip, p_cols['CAPACITY']).value = w['c_cap']
                        
                        # Mech is Zero for composers
                        if p_cols['MECH_OWN']: ws_ip.cell(write_row_ip, p_cols['MECH_OWN']).value = 0
                        if p_cols['MECH_COLL']: ws_ip.cell(write_row_ip, p_cols['MECH_COLL']).value = 0
                        
                        # Perf is 50% of Share
                        if p_cols['PERF_OWN']: ws_ip.cell(write_row_ip, p_cols['PERF_OWN']).value = comp_perf
                        if p_cols['PERF_COLL']: ws_ip.cell(write_row_ip, p_cols['PERF_COLL']).value = comp_perf

                    write_row_ip += 1

            # ---------------------------
            # 2. ALTERNATE TITLES LOGIC (Unchanged)
            # ---------------------------
            
            key = (w_title_norm, w_foreign_norm)
            alt_titles = jot_alt_lookup.get(key)

            if alt_titles:
                for alt in alt_titles:
                    capitalized_alt = str(alt).strip().title()
                    
                    # --- Alternate Titles Sheet Columns (Target) ---
                    col_alt_work_id = alt_headers.get("WORK ID") or alt_headers.get("ID")
                    col_alt_title = alt_headers.get("WORK TITLE")
                    col_alt_main = alt_headers.get("WORK MAIN IDENTIFIER")
                    col_alt_tune = alt_headers.get("WORK TUNECODE")
                    col_alt_alt = alt_headers.get("ALTERNATE TITLE")
                    col_alt_lang = alt_headers.get("LANGUAGE")

                    if col_alt_work_id: ws_alt.cell(write_row_alt, col_alt_work_id).value = w_work_id_val
                    if col_alt_title: ws_alt.cell(write_row_alt, col_alt_title).value = w_title_val
                    if col_alt_main: ws_alt.cell(write_row_alt, col_alt_main).value = w_main_id_val if w_main_id_val else ""
                    if col_alt_tune: ws_alt.cell(write_row_alt, col_alt_tune).value = w_tunecode_val if w_tunecode_val else ""
                    
                    if col_alt_alt: ws_alt.cell(write_row_alt, col_alt_alt).value = capitalized_alt 
                    if col_alt_lang: ws_alt.cell(write_row_alt, col_alt_lang).value = w_lang_val

                    write_row_alt += 1

        # ---- Save ----
        out = io.BytesIO()
        wb.save(out) 
        out.seek(0)
        return out
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return f"Processing Failed in step2.py: {e.__class__.__name__}: {e}"